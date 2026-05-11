import streamlit as st
import pdfplumber
import re
import zipfile
import os
import pandas as pd
from io import BytesIO
from tempfile import TemporaryDirectory
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def app():
    st.markdown(
        "<h1 style='white-space: nowrap; font-size: 44px;'>📄 Leitor PDF | NF3e - Energia Elétrica</h1>",
        unsafe_allow_html=True
    )
    st.markdown("""
Essa ferramenta permite extrair automaticamente os principais dados fiscais de contas de energia elétrica no formato PDF (modelo NF3e). 
Você pode enviar um ou vários arquivos `.pdf`, ou um `.zip` contendo múltiplos PDFs. 
O sistema identifica e organiza automaticamente as informações extraídas dos PDFs, gerando uma planilha Excel pronta para conferência.
""")

    # Upload de arquivos (agora DENTRO da função)
    uploaded_files = st.file_uploader("Envie múltiplos arquivos (.pdf) ou um (.zip) contendo vários PDFs.", type=["pdf", "zip"], accept_multiple_files=True)

    # === Função: Extrair dados de um PDF ===
    def extrair_dados_pdf(file):
        with pdfplumber.open(file) as pdf:
            texto = ""
            for page in pdf.pages:
                texto += page.extract_text() + "\n"

        def buscar(regex, flags=0):
            match = re.search(regex, texto, flags)
            return match.group(1).strip() if match else None

        uc_conteudo = buscar(r'\n(?:JAN|FEV|MAR|ABR|MAI|JUN|JUL|AGO|SET|OUT|NOV|DEZ)/\d{4}\s*\n?(\d{9,12})')
        uc_arquivo = os.path.splitext(os.path.basename(file.name))[0]

        if uc_conteudo and uc_conteudo != uc_arquivo:
            st.warning(f"⚠️ UC divergente: conteúdo do PDF → {uc_conteudo}, nome do arquivo → {uc_arquivo}")

        valor = buscar(r'R\$\*{5,}(\d{1,3},\d{2})') or buscar(r'(\d{1,3},\d{2})\nO Pagamento poderá ser realizado')

        return {
            "Nota Fiscal": buscar(r'NOTA FISCAL Nº (\d+)'),
            "Série": buscar(r'NOTA FISCAL Nº \d+\s*-\s*SÉRIE\s*(\S+)'),
            "CNPJ": buscar(r'CNPJ/CPF:\s*([\d./-]+)'),
            "Valor (R$)": valor,
            "Data de Emissão": buscar(r'DATA DE EMISSÃO:\s*(\d{2}/\d{2}/\d{4})'),
            "Nome do Destinatário": buscar(r'^\s*(ROMA HOTEIS.*FILIAL VILLAS)', re.MULTILINE),
            "Protocolo de Autorização": buscar(r'Protocolo de autorização:\s*(.*?)\s*-'),
            "Unidade Consumidora": uc_arquivo,
            "Chave de Acesso": buscar(r'chave de acesso:\s*([\d]+)')
        }

    def processar_arquivos(files):
        dados_extraidos = []
        for file in files:
            if file.name.endswith(".pdf"):
                dados = extrair_dados_pdf(file)
                dados_extraidos.append(dados)
            elif file.name.endswith(".zip"):
                with TemporaryDirectory() as tmpdir:
                    zip_path = os.path.join(tmpdir, file.name)
                    with open(zip_path, "wb") as f:
                        f.write(file.read())
                    with zipfile.ZipFile(zip_path, "r") as zip_ref:
                        zip_ref.extractall(tmpdir)
                        for nome_arquivo in zip_ref.namelist():
                            if nome_arquivo.endswith(".pdf"):
                                caminho_pdf = os.path.join(tmpdir, nome_arquivo)
                                with open(caminho_pdf, "rb") as f_pdf:
                                    dados = extrair_dados_pdf(f_pdf)
                                    dados_extraidos.append(dados)
        return dados_extraidos

    if uploaded_files:
        with st.spinner("⏳ Extraindo dados dos arquivos..."):
            resultado = processar_arquivos(uploaded_files)
            df_resultado = pd.DataFrame(resultado)

        if "Valor (R$)" in df_resultado.columns:
            df_resultado["Valor (R$)"] = df_resultado["Valor (R$)"].str.replace(",", ".").astype(float)
        else:
            st.warning("Coluna 'Valor (R$)' não encontrada nos dados extraídos. Verifique o arquivo enviado.")
        st.success("✅ Dados extraídos com sucesso!")
        st.dataframe(df_resultado)

        output = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Notas Fiscais"
        ws.freeze_panes = "A2"

        for r in dataframe_to_rows(df_resultado, index=False, header=True):
            ws.append(r)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="000000")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "Chave de Acesso":
                chave_col_idx = idx
                break
        for row in ws.iter_rows(min_row=2, min_col=chave_col_idx, max_col=chave_col_idx):
            for cell in row:
                cell.number_format = "@"

        ws.sheet_view.showGridLines = False
        wb.save(output)
        st.download_button("📥 Baixar Planilha", data=output.getvalue(), file_name="dados_nfe3.xlsx")
