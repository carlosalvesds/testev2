# ferramentas/leitor_rt.py

import streamlit as st
import pandas as pd
import zipfile
import os
import tempfile
import xml.etree.ElementTree as ET
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from concurrent.futures import ThreadPoolExecutor

def map_crt(crt):
    return {
        '1': 'Simples Nacional',
        '2': 'Simples Nacional, excesso sublimite de receita bruta',
        '3': 'Regime Normal',
        '4': 'Microempreendedor Individual'
    }.get(crt, 'Não identificado')

def format_cnpj_cpf(doc):
    doc = re.sub(r'\D', '', doc)
    if len(doc) == 14:
        return f"{doc[:2]}.{doc[2:5]}.{doc[5:8]}/{doc[8:12]}-{doc[12:]}"
    elif len(doc) == 11:
        return f"{doc[:3]}.{doc[3:6]}.{doc[6:9]}-{doc[9:]}"
    return doc

def process_xml_file(xml_file, ns):
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
        cnpj = root.find('.//ns:emit/ns:CNPJ', ns)
        nome = root.find('.//ns:emit/ns:xNome', ns)
        crt = root.find('.//ns:emit/ns:CRT', ns)
        if cnpj is not None and nome is not None and crt is not None:
            return {
                'CNPJ': cnpj.text,
                'Nome da Empresa': nome.text,
                'Regime Tributário': map_crt(crt.text)
            }
    except:
        return None

def gerar_excel_formatado(df, caminho_saida, total_lidos, removidos, total_extraidos):
    df['CNPJ'] = df['CNPJ'].apply(format_cnpj_cpf)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Regime_Tributário"
    ws.sheet_view.showGridLines = False

    ws['A1'] = f"Total de XMLs lidos: {total_lidos}"
    ws['B1'] = f"Duplicidades removidas: {removidos}"
    ws['C1'] = f"Total após exclusão: {total_extraidos}"
    ws.merge_cells('A1:C1')
    ws['A1'].font = Font(bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 3):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 3:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_cells in ws.iter_cols(min_row=3):
        col_letter = col_cells[0].column_letter
        max_length = 0
        for cell in col_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                continue
        adjusted_width = max_length * 1.2 + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(caminho_saida)

def app():
    st.title("📁 XML NF-e | Regime Tributário")
    st.markdown("""
Essa ferramenta analisa arquivos XML de NF-e para identificar automaticamente o regime tributário de cada fornecedor, 
organizando os dados em uma planilha Excel. 
Suporta o processamento de grandes volumes de XMLs, garantindo eficiência mesmo em operações com milhares de documentos.
""")

    uploaded_file = st.file_uploader(" Envie o arquivo (.zip) contendo os XMLs das NF-e", type="zip")

    if uploaded_file is not None:
        with tempfile.TemporaryDirectory() as tmpdir:
            zip_path = os.path.join(tmpdir, "arquivo.zip")
            with open(zip_path, "wb") as f:
                f.write(uploaded_file.getvalue())
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(tmpdir)

            xml_files = []
            for root_dir, dirs, files in os.walk(tmpdir):
                for file in files:
                    if file.lower().endswith('.xml'):
                        xml_files.append(os.path.join(root_dir, file))

            ns = {'ns': 'http://www.portalfiscal.inf.br/nfe'}
            with ThreadPoolExecutor() as executor:
                resultados = list(executor.map(lambda x: process_xml_file(x, ns), xml_files))

            resultados_filtrados = [r for r in resultados if r]
            total_lidos = len(resultados)
            df = pd.DataFrame(resultados_filtrados)
            total_antes = len(df)
            df = df.drop_duplicates()
            total_depois = len(df)
            removidos = total_antes - total_depois

            if not df.empty:
                st.success(f"✅ {total_depois} XMLs extraídos com sucesso.")
                st.info(f"📄 Total de XMLs lidos: {total_lidos}")
                st.info(f"♻️ Duplicidades removidas: {removidos}")
                st.info(f"📊 Total após exclusão: {total_depois}")
                st.dataframe(df)

                excel_path = os.path.join(tmpdir, "Regime_Tributario_Formatado.xlsx")
                gerar_excel_formatado(df, excel_path, total_lidos, removidos, total_depois)

                with open(excel_path, "rb") as f:
                    st.download_button(
                        label="📥 Baixar Planilha",
                        data=f,
                        file_name="Regime_Tributario.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                st.warning("⚠️ Nenhum dado foi extraído. Verifique se os arquivos XML estão no padrão correto.")
